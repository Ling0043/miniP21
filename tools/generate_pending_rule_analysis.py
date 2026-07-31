from pathlib import Path
import json
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "dataset" / "P21_rule_wait_to_do.xlsx"
OUTPUT = ROOT / "dataset" / "P21_rule_wait_to_do_analysis.xlsx"

HEADERS = [
    "ruleId", "message", "description", "domains",
    "reuse_existing_coreFunction(是/否)", "suggested_coreFunction",
    "suggested_targetVariable", "suggested_ruleParams",
    "cross_domain_check(是/否)", "analysis_notes",
]


def params(**kwargs):
    return json.dumps(kwargs, ensure_ascii=False, separators=(",", ":"))


def existing(core, target, rule_params, note):
    return ("是", core, target, rule_params, "否", note)


REUSE = {
    "SD0041": existing("check_be_null_cond", "--OCCUR", params(
        cond_vars=["--PRESP"], cond_ops=["missing"], cond_vals=[""], logic_op="AND"
    ), "现有条件必空函数可表达 PRESP 缺失时 OCCUR 不应填报。"),
    "SD0047": existing("check_missing_cond", "--ORRES", params(
        cond_vars=["--STAT", "--DRVFL"], cond_ops=["not_in", "not_in"],
        cond_vals=["NOT DONE,__MISSING__", "Y,__MISSING__"], logic_op="AND"
    ), "复用条件必填函数；STAT 非 NOT DONE 且 DRVFL 非 Y 时 ORRES 必填。"),
    "SD0088": existing("check_missing_cond", "RFENDTC", params(
        cond_vars=["ARMCD"], cond_ops=["not_in"],
        cond_vals=["SCRNFAIL,NOTASSGN,__MISSING__"], logic_op="AND"
    ), "复用随机受试者参考结束日期必填的条件模板。"),
    "SD1072": existing("check_missing_cond", "IDVAR", params(
        cond_vars=["RDOMAIN"], cond_ops=["not_in"], cond_vals=["DM,__MISSING__"], logic_op="AND"
    ), "RDOMAIN 已填且非 DM 时 IDVAR 必填，现有 not_in 条件可表达。"),
    "SD1263": existing("check_missing_cond", "MBRESCAT", params(
        cond_vars=["MBTESTCD", "MBSTRESC", "MBMETHOD"],
        cond_ops=["equal", "not_in", "not_in"],
        cond_vals=["ORGANISM", "NO GROWTH,__MISSING__", "GRAM STAIN,__MISSING__"], logic_op="AND"
    ), "ORGANISM 且非例外结果/方法时 MBRESCAT 必填，现有多条件必填函数可表达。"),
    "SD1272": existing("check_be_null_cond", "--TESTCD", params(
        cond_vars=["--TESTCD"], cond_ops=["equal"], cond_vals=["OTHER"], logic_op="AND"
    ), "复用条件必空函数禁止 TESTCD=OTHER。"),
    "SD1273": existing("check_be_null_cond", "--TRT", params(
        cond_vars=["--TRT"], cond_ops=["equal"], cond_vals=["OTHER"], logic_op="AND"
    ), "复用条件必空函数禁止 TRT=OTHER。"),
    "SD1274": existing("check_be_null_cond", "--TERM", params(
        cond_vars=["--TERM"], cond_ops=["equal"], cond_vals=["OTHER"], logic_op="AND"
    ), "复用条件必空函数禁止 TERM=OTHER。"),
    "SD1275": existing("check_be_null_cond", "--TESTCD", params(
        cond_vars=["--TESTCD"], cond_ops=["equal"], cond_vals=["MULTIPLE"], logic_op="AND"
    ), "复用条件必空函数禁止 TESTCD=MULTIPLE。"),
    "SD1276": existing("check_be_null_cond", "--TRT", params(
        cond_vars=["--TRT"], cond_ops=["equal"], cond_vals=["MULTIPLE"], logic_op="AND"
    ), "复用条件必空函数禁止 TRT=MULTIPLE。"),
    "SD1277": existing("check_be_null_cond", "--TERM", params(
        cond_vars=["--TERM"], cond_ops=["equal"], cond_vals=["MULTIPLE"], logic_op="AND"
    ), "复用条件必空函数禁止 TERM=MULTIPLE。"),
    "SD1279": existing("check_missing_cond", "ECDOSTXT", params(
        cond_vars=["ECDOSE", "ECOCCUR", "ECSTAT"],
        cond_ops=["missing", "not_in", "not_in"],
        cond_vals=["", "N,__MISSING__", "NOT DONE,__MISSING__"], logic_op="AND"
    ), "ECDOSE 缺失且并非未发生/未完成时 ECDOSTXT 必填。"),
    "SD1307": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="TDIGRP", cond_param="HLTSUBJI", cond_val=["N"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1308": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="CURTRT", cond_param="ADDON", cond_val=["Y"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1309": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="TRT", cond_param="STYPE", cond_val=["INTERVENTIONAL"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1310": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="INTMODEL", cond_param="STYPE", cond_val=["INTERVENTIONAL"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1311": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="INTTYPE", cond_param="STYPE", cond_val=["INTERVENTIONAL"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1312": existing("check_ts_param_required", "TSPARMCD", params(
        param_code="PCLAS", cond_param="STYPE", cond_val=["INTERVENTIONAL"], val_var="TSVAL"
    ), "与现有 TS 条件必填模板同构。"),
    "SD1326": existing("check_be_null_cond", "RDEVID", params(
        cond_vars=["RSUBJID"], cond_ops=["non_missing"], cond_vals=[""], logic_op="AND"
    ), "RSUBJID 已填时 RDEVID 必须为空，可复用条件必空函数。"),
    "SD1332": existing("check_be_null_cond", "AEENDTC", params(
        cond_vars=["AEOUT"], cond_ops=["equal"], cond_vals=["NOT RECOVERED/NOT RESOLVED"], logic_op="AND"
    ), "AEOUT 为未恢复/未解决时 AEENDTC 必须为空。"),
    "SD1360": existing("check_be_null_cond", "ARMNRS", params(
        cond_vars=["ARMCD", "ACTARMCD"], cond_ops=["non_missing", "non_missing"],
        cond_vals=["", ""], logic_op="AND"
    ), "ARMCD 与 ACTARMCD 均已填时 ARMNRS 必须为空。"),
    "SD1366": existing("check_be_null_cond", "RFENDTC", params(
        cond_vars=["ARMNRS"], cond_ops=["not_in"],
        cond_vals=["UNPLANNED TREATMENT,__MISSING__"], logic_op="AND"
    ), "ARMNRS 已填且非 UNPLANNED TREATMENT 时 RFENDTC 必须为空。"),
    "SD1371": existing("check_missing_cond", "--ORRES", params(
        cond_vars=["--LOBXFL", "--DRVFL"], cond_ops=["equal", "not_in"],
        cond_vals=["Y", "Y,__MISSING__"], logic_op="AND"
    ), "LOBXFL=Y 且非派生记录时 ORRES 必填。"),
    "SD1373": existing("check_be_null_cond", "RFSTDTC", params(
        cond_vars=["ARMNRS"], cond_ops=["not_in"],
        cond_vals=["UNPLANNED TREATMENT,__MISSING__"], logic_op="AND"
    ), "ARMNRS 已填且非 UNPLANNED TREATMENT 时 RFSTDTC 必须为空。"),
    "SD1440": existing("check_consistent", "--DECOD", params(group_vars="--TERM"), "同一 TERM 内 DECOD 一致性可复用现有分组一致性函数。"),
    "SD1441": existing("check_consistent", "--HLT", params(group_vars="--TERM"), "同一 TERM 内 HLT 一致性可复用现有分组一致性函数。"),
    "SD1442": existing("check_consistent", "--HLGT", params(group_vars="--TERM"), "同一 TERM 内 HLGT 一致性可复用现有分组一致性函数。"),
    "SD1444": existing("check_consistent", "--SOC", params(group_vars="--TERM"), "同一 TERM 内 SOC 一致性可复用现有分组一致性函数。"),
    "SD2244": existing("check_expected_cond", "TSVCDREF", params(
        cond_vars=["TSPARMCD"], cond_vals=["FCNTRY"], expected_val="GENC"
    ), "与现有 TSVCDREF 条件固定值规则同构。"),
    "SD2247A": existing("check_iso8601_cond", "TSVAL", params(
        cond_var="TSPARMCD", cond_val="SSTDTC", type="date"
    ), "与历史 SD2247 的 SSTDTC 日期格式配置一致。"),
}

CROSS_DOMAIN = {
    "CT2001", "CT2002", "CT2003", "CT2004", "CT2005", "CT2006",
    "SD0006", "SD0064", "SD0065", "SD0066", "SD0067", "SD0068", "SD0070", "SD0071",
    "SD0072", "SD0072A", "SD0075", "SD0077", "SD0079", "SD0080", "SD0082",
    "SD1005", "SD1012", "SD1014", "SD1015", "SD1016", "SD1017", "SD1018", "SD1023",
    "SD1030", "SD1031", "SD1084", "SD1085", "SD1086", "SD1088", "SD1089", "SD1090",
    "SD1092", "SD1093", "SD1094", "SD1144", "SD1202", "SD1203", "SD1204", "SD1205",
    "SD1206", "SD1207", "SD1233", "SD1240", "SD1252", "SD1253", "SD1254", "SD1255",
    "SD1256", "SD1261", "SD1262", "SD1316", "SD1317", "SD1318", "SD1319", "SD1322",
    "SD1336", "SD1337", "SD1338", "SD1339", "SD1340", "SD1347", "SD1354", "SD1374",
    "SD1377", "SD1378", "SD1379", "SD1446", "SD1477", "SD1478", "SD1482", "SD1485",
    "SD2002", "SD2003", "SD2006", "SD2236", "SD2237", "SD2250", "SD2251", "SD2252",
    "SD2253", "SD2254", "SD2255", "SD2257", "SD2258", "SD2259", "SD2260", "SD2261",
    "SD2262", "SD2263", "SD2264", "SD2265", "SD2267", "SD2268", "SD2269", "SD2283", "SD2288",
}

INAPPROPRIATE = {
    "SD0062", "SD0095", "SD1071", "SD1079", "SD1095", "SD1116", "SD1142", "SD1457",
    "SD1474", "SD1475", "SD9999",
}


def extract_target(message, description):
    candidates = re.findall(r"--[A-Z0-9]+|(?<![A-Z0-9])(?:[A-Z][A-Z0-9]{1,7})(?![A-Z0-9])", message)
    if not candidates:
        candidates = re.findall(r"--[A-Z0-9]+|(?<![A-Z0-9])(?:[A-Z][A-Z0-9]{1,7})(?![A-Z0-9])", description)
    ignored = {"SDTM", "IG", "CDISC", "ALL", "ISO", "SAS", "FDA", "PMDA", "SEND", "CT"}
    targets = []
    for candidate in candidates:
        if candidate not in ignored and candidate not in targets:
            targets.append(candidate)
    return ",".join(targets[:4])


def default_analysis(rule_id, message, description):
    target = extract_target(message, description)
    if rule_id in INAPPROPRIATE:
        return ("否", "inappropriate", "", "", "否", "依赖文件、数据集元数据或数据集类别，当前单个 domain data.frame 校验函数不适用。")
    if rule_id in CROSS_DOMAIN:
        return ("否", f"check_{rule_id.lower()}", target, "{}", "是", "需要读取其他 domain 或外部 CT/字典；现有单域调度框架暂不支持。")
    return ("否", f"check_{rule_id.lower()}", target, "{}", "否", "现有函数无法完整表达该规则语义；建议新增该规则专用单域函数，并在实现时定义参数契约。")


def main():
    source_book = load_workbook(SOURCE, read_only=True, data_only=True)
    source_sheet = source_book.active
    rows = list(source_sheet.iter_rows(min_row=2, values_only=True))
    if len(rows) != 250:
        raise ValueError(f"Expected 250 pending rules, found {len(rows)}")

    book = Workbook()
    sheet = book.active
    sheet.title = "rule_analysis"
    sheet.append(HEADERS)

    for rule_id, message, description, domains in rows:
        result = REUSE.get(rule_id, default_analysis(rule_id, message, description))
        sheet.append([rule_id, message, description, domains, *result])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 42, 92, 46, 18, 28, 30, 68, 16, 58]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    book.save(OUTPUT)
    print(f"Wrote {len(rows)} rules to {OUTPUT}")


if __name__ == "__main__":
    main()